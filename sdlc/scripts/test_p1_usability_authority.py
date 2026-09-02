#!/usr/bin/env python3
"""P1 self-test for decision authority, schema-safe overlay, Worklist sync and knowledge promotion."""
from __future__ import annotations

import tempfile
from pathlib import Path
import sys
import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SCRIPTS = ROOT / "sdlc" / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import build_project_decision_registry as decisions  # noqa: E402
import resolve_project_overlay as overlays  # noqa: E402
import sync_worklist as worklist  # noqa: E402
import promote_knowledge as knowledge  # noqa: E402
import validate_contract_authority as authority  # noqa: E402


def load(rel: str):
    return yaml.safe_load((ROOT / rel).read_text(encoding="utf-8")) or {}


def test_decision_registry():
    bootstrap = {"project_bootstrap": {"project_id": "GF-1", "resolved_mode": "GREENFIELD", "technology_decisions": []}}
    config = load("sdlc/config/project-decisions.yaml")
    result = decisions.build(bootstrap, config)
    root = result["project_decisions"]
    assert set(root["decisions"]) == set(config["decisions"])
    assert any(x["action"] == "source.write" for x in root["action_blockers"])
    assert all(x["state"] in {"OPEN", "NOT_APPLICABLE"} for x in root["decisions"].values())
    assert decisions.validate(result, config) == []


def active_overlay(target_key: str, current, new):
    return {"overlay": {
        "overlay_id": "OVL-1",
        "scope": {"project_id": "P-1", "domain": "", "applies_to": []},
        "state": "ACTIVE", "revision": 1,
        "trigger": {"type": "REQUIRED_ARTIFACT_OR_STAGE_DIFFERS", "reason": "customer project requires this value", "detected_at_stage": "BOOTSTRAP", "detected_by": "selftest"},
        "basis": {"truth_state": "GIVEN", "evidence_ids": [], "source_refs": [], "project_fact": "customer supplied setting"},
        "change": {"target_key": target_key, "core_or_profile_value": current, "project_value": new, "rationale": "selftest"},
        "safety": {"copies_core_truth": False, "sample_specific_only": False},
        "lifecycle": {"activated_by": "owner", "activated_at": "2026-09-02T00:00:00Z"},
    }}


def test_overlay_schema_safe():
    profile = {"project": {"name": "P", "mode": "AUTO"}, "artifacts": {"profile": "STANDARD"}, "providers": {"registry": "registry.yaml"}, "customization": {"overlays": []}}
    schema = load("sdlc/config/overlay-schema.yaml")
    resolved, report = overlays.resolve(profile, [("ok", active_overlay("artifacts.profile", "STANDARD", "LITE"))], schema)
    assert resolved["artifacts"]["profile"] == "LITE" and len(report["applied"]) == 1
    try:
        overlays.resolve(profile, [("bad", active_overlay("artifacts.profle", None, "LITE"))], schema)
    except ValueError as exc:
        assert "unknown target_key" in str(exc)
    else:
        raise AssertionError("unknown overlay key must be denied")
    try:
        overlays.resolve(profile, [("type", active_overlay("project.name", "P", ["bad"]))], schema)
    except ValueError as exc:
        assert "type" in str(exc).lower()
    else:
        raise AssertionError("overlay type change must be denied")


def seed_item():
    return {
        "work_item_id": "TASK-1", "parent_id": "", "requirement_id": "RQ-1", "item_type": "TASK", "name": "Selftest", "stage": "DEVELOPMENT",
        "status": "READY", "quality": "OK", "validity": "CURRENT", "assignee": "", "planned_start": "", "planned_end": "", "estimated_effort": "",
        "actual_start": "", "actual_end": "", "actual_effort": "", "dependency_ids": "", "program_ids": "PGM-1", "acceptance_test_ids": "AC-1",
        "alerts": "", "updated_at": "2026-09-02T00:00:00Z", "revision": "1", "note": "",
    }


def write_canonical(path: Path, items: list[dict]):
    path.write_text(yaml.safe_dump({"schema_version": 1, "artifact_type": "WORKLIST_CANONICAL", "worklist_canonical": {"items": items}}, allow_unicode=True, sort_keys=False), encoding="utf-8")


def test_worklist_roundtrip_and_conflict():
    cfg = load("sdlc/config/worklist-columns.yaml")
    with tempfile.TemporaryDirectory() as td:
        root = Path(td); canonical = root / "canonical.yaml"; md = root / "worklist.md"; xlsx = root / "worklist.xlsx"
        write_canonical(canonical, [seed_item()])
        rc = worklist.main_with_args(["--canonical", str(canonical), "--md", str(md), "--xlsx", str(xlsx), "--config", str(ROOT / "sdlc/config/worklist-columns.yaml")]) if hasattr(worklist, "main_with_args") else None
        if rc is None:
            merged, conflicts, changed = worklist.merge([], [], worklist.canonical_rows(canonical), cfg)
            assert not conflicts and not changed
            worklist.write_md(md, merged, cfg); worklist.write_xlsx(xlsx, merged, cfg)
        assert md.exists() and xlsx.exists()

        rows = worklist.parse_md(md, cfg); rows[0]["status"] = "IN_PROGRESS"; rows[0]["revision"] = "2"; worklist.write_md(md, rows, cfg)
        md_rows = worklist.parse_md(md, cfg); xls_rows = worklist.parse_xlsx(xlsx, cfg); can_rows = worklist.canonical_rows(canonical)
        merged, conflicts, changed = worklist.merge(md_rows, xls_rows, can_rows, cfg)
        assert not conflicts and changed == ["TASK-1"]
        write_canonical(canonical, merged); worklist.write_md(md, merged, cfg); worklist.write_xlsx(xlsx, merged, cfg)
        assert worklist.canonical_rows(canonical)[0]["status"] == "IN_PROGRESS"

        bad_md = worklist.parse_md(md, cfg); bad_md[0]["status"] = "DONE"; worklist.write_md(md, bad_md, cfg)
        _, unversioned, _ = worklist.merge(worklist.parse_md(md, cfg), worklist.parse_xlsx(xlsx, cfg), worklist.canonical_rows(canonical), cfg)
        assert unversioned and unversioned[0]["code"] == "UNVERSIONED_EDIT"
        worklist.write_md(md, worklist.canonical_rows(canonical), cfg)

        md_rows = worklist.parse_md(md, cfg); md_rows[0]["status"] = "DONE"; md_rows[0]["revision"] = "3"; worklist.write_md(md, md_rows, cfg)
        wb = load_workbook(xlsx); ws = wb.active
        labels = [c["label"] for c in cfg["columns"]]; status_col = labels.index("상태") + 1; rev_col = labels.index("변경버전") + 1
        ws.cell(2, status_col).value = "BLOCKED"; ws.cell(2, rev_col).value = 3; wb.save(xlsx); wb.close()
        _, conflict, _ = worklist.merge(worklist.parse_md(md, cfg), worklist.parse_xlsx(xlsx, cfg), worklist.canonical_rows(canonical), cfg)
        assert conflict and conflict[0]["code"] == "SYNC_CONFLICT"
        assert worklist.canonical_rows(canonical)[0]["revision"] == "2"


def reviewed_candidate(truth: str, human_confirmation: bool):
    return {"knowledge_candidate": {
        "knowledge_id": "K-BR-1", "project_id": "P-1", "type": "BUSINESS_RULE", "title": "Rule", "statement": "candidate statement",
        "truth_state": truth, "promotion_state": "CANDIDATE", "revision": 1, "applies_to_ids": ["RQ-1"],
        "provenance": {"evidence_ids": ["EV-1"], "source_refs": [], "source_revision": "rev-1"},
        "review": {"required": True, "decision": "CONFIRM", "human_confirmation": human_confirmation, "reviewed_by": "owner", "reviewed_at": "2026-09-02T00:00:00Z", "decision_basis": "customer confirmation"},
        "relations": {"supports": [], "conflicts_with": [], "supersedes": []},
    }}


def test_knowledge_promotion_truth_guard():
    config = load("sdlc/config/knowledge-promotion.yaml")
    _, denied = knowledge.promote(reviewed_candidate("OBSERVED", False), {}, config)
    assert denied["state"] == "DENIED" and any("P1KP-012" in x for x in denied["errors"])
    registry, promoted = knowledge.promote(reviewed_candidate("OBSERVED", True), {}, config)
    assert promoted["state"] == "PROMOTED" and promoted["truth_state"] == "CONFIRMED"
    assert registry["knowledge_registry"]["entries"][0]["truth_state_before_review"] == "OBSERVED"
    assert registry["knowledge_registry"]["canonical_publish_automatic"] is False
    registry2, idem = knowledge.promote(reviewed_candidate("OBSERVED", True), registry, config)
    assert idem["state"] == "IDEMPOTENT" and registry2 == registry


def test_authority_index_shape():
    config = load("sdlc/config/contract-authority.yaml")
    with tempfile.TemporaryDirectory() as td:
        fake = Path(td)
        for spec in (config.get("authorities") or {}).values():
            if spec.get("generated"): continue
            path = fake / spec["path"]; path.parent.mkdir(parents=True, exist_ok=True); path.write_text("x", encoding="utf-8")
        assert authority.validate(fake, config) == []


def test_guide_numbering_and_design_boundary():
    guides = ROOT / "sdlc" / "guides"
    numbered = sorted((x.name for x in guides.glob("[0-9][0-9]_*.md")), key=lambda name: int(name.split("_", 1)[0]))
    numbers = [name.split("_", 1)[0] for name in numbered]
    assert len(numbers) == len(set(numbers)), f"duplicate guide number: {numbered}"
    assert numbered == [
        "01_SDLC_전체가이드.md",
        "02_SKILL_사용가이드.md",
        "03_TEMPLATE_산출물가이드.md",
        "04_PROVIDER_RUNTIME_사용가이드.md",
        "05_HARNESS_커스터마이징가이드.md",
    ]
    config = load("sdlc/config/contract-authority.yaml")
    roots = set(config.get("non_authoritative_roots") or [])
    assert "sdlc/design" in roots and "sdlc/guides" in roots
    assert (ROOT / "sdlc/design/README.md").is_file()
    assert (ROOT / "sdlc/guides/README.md").is_file()


def main() -> int:
    tests = [test_decision_registry, test_overlay_schema_safe, test_worklist_roundtrip_and_conflict, test_knowledge_promotion_truth_guard, test_authority_index_shape, test_guide_numbering_and_design_boundary]
    for test in tests:
        test(); print(f"PASS {test.__name__}")
    print(f"PASS P1 usability/authority tests={len(tests)}")
    return 0


if __name__ == "__main__": raise SystemExit(main())
